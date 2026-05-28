#!/usr/bin/env python3
"""Multi-format processor: PDF, DOCX, XLSX, archives, audio/video transcription."""
import os, re, json, shutil, tempfile, hashlib, time
import requests

FINNA_KEY = os.environ.get("FINNA_KEY", "app-ULzJbc3OaIN50mZVSU7sAa97")
FINNA_BASE = "https://www.finna.com.cn/v1"
ASR_KEY = os.environ.get("ASR_KEY", "app-1t7UfcSJIB5r0N20vbDaid9A")

ARCHIVE_EXTS = {'.zip', '.tar.gz', '.tgz', '.tar.bz2', '.tbz2', '.tar.xz', '.tar'}
MEDIA_EXTS = {'.mp3', '.wav', '.m4a', '.aac', '.ogg', '.flac', '.wma', '.amr', '.opus', '.wav'}
VIDEO_EXTS = {'.mp4', '.avi', '.mov', '.mkv', '.webm', '.flv', '.wmv', '.mpeg', '.mpg', '.3gp'}

def get_ext(filename):
    """Get file extension, handling compound extensions."""
    f = filename.lower()
    if f.endswith('.tar.gz'): return '.tar.gz'
    if f.endswith('.tar.bz2'): return '.tar.bz2'
    if f.endswith('.tar.xz'): return '.tar.xz'
    if f.endswith('.tgz'): return '.tar.gz'
    if f.endswith('.tbz2'): return '.tar.bz2'
    return os.path.splitext(f)[1]

def is_archive(filename):
    return get_ext(filename) in ARCHIVE_EXTS

def is_media(filename):
    ext = get_ext(filename)
    return ext in MEDIA_EXTS or ext in VIDEO_EXTS

def is_video(filename):
    return get_ext(filename) in VIDEO_EXTS


# ─── Archive Extraction ───
def extract_archive(filepath, filename):
    """Extract archive to temp dir. Returns list of {filepath, filename, ext}."""
    ext = get_ext(filename)
    tmpdir = tempfile.mkdtemp(prefix='nb-archive-')
    extracted = []

    try:
        if ext == '.zip':
            import zipfile
            with zipfile.ZipFile(filepath, 'r') as zf:
                for member in zf.namelist():
                    if member.endswith('/') or member.startswith('__MACOSX'):
                        continue
                    # Safe extraction
                    zf.extract(member, tmpdir)
                    full_path = os.path.join(tmpdir, member)
                    if os.path.isfile(full_path):
                        extracted.append({
                            'filepath': full_path,
                            'filename': os.path.basename(member),
                            'ext': get_ext(os.path.basename(member))
                        })
        elif ext in ('.tar.gz', '.tar.bz2', '.tar.xz', '.tar'):
            import tarfile
            mode_map = {'.tar.gz': 'r:gz', '.tar.bz2': 'r:bz2', '.tar.xz': 'r:xz', '.tar': 'r'}
            mode = mode_map.get(ext, 'r:*')
            with tarfile.open(filepath, mode) as tf:
                for member in tf.getmembers():
                    if not member.isfile():
                        continue
                    # Safe extraction (no path traversal)
                    member.name = os.path.basename(member.name)
                    tf.extract(member, tmpdir, filter='data')
                    full_path = os.path.join(tmpdir, member.name)
                    if os.path.isfile(full_path):
                        extracted.append({
                            'filepath': full_path,
                            'filename': member.name,
                            'ext': get_ext(member.name)
                        })
    except Exception as e:
        print(f"[Archive] Extraction error: {e}")
        # Return what we have
    return extracted, tmpdir


# ─── ASR Transcription ───
def transcribe_media(filepath, filename):
    """Transcribe audio/video via FinnA ASR API. Returns text string."""
    try:
        with open(filepath, 'rb') as f:
            resp = requests.post(
                f"{FINNA_BASE}/audio/transcriptions",
                headers={"Authorization": f"Bearer {ASR_KEY}"},
                files={"file": (filename, f)},
                data={"model": "qwen3-asr-flash"},
                timeout=300,
                stream=True
            )

        if resp.status_code != 200:
            return f"[ASR Error: HTTP {resp.status_code}]"

        text = ""
        for line in resp.iter_lines(decode_unicode=True):
            if not line or not line.startswith('data: '):
                continue
            data_str = line[6:]
            if data_str == '[DONE]':
                break
            try:
                data = json.loads(data_str)
                if data.get('event') == 'transcript.text.done':
                    text = data.get('text', '')
                    break
            except json.JSONDecodeError:
                continue

        return text if text else "[ASR: No transcription result]"

    except requests.exceptions.Timeout:
        return "[ASR Error: Timeout]"
    except Exception as e:
        return f"[ASR Error: {str(e)}]"


# ─── XLSX Extraction ───
def extract_xlsx(filepath, filename):
    """Extract text from Excel file with table formatting."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"\n## Sheet: {sheet_name}\n")

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # Detect header row
            max_cols = max(len([c for c in row if c is not None]) for row in rows[:5])

            for row_idx, row in enumerate(rows):
                cells = [str(c) if c is not None else '' for c in row[:max_cols]]
                if not any(cells):
                    continue

                line = ' | '.join(cells)
                # Mark first row as header
                if row_idx == 0:
                    parts.append(f"| {' | '.join(cells)} |")
                    parts.append(f"|{'---|' * max_cols}")
                else:
                    parts.append(f"| {' | '.join(cells)} |")

            parts.append("")

        wb.close()
        return '\n'.join(parts) if parts else "[XLSX: Empty workbook]"

    except Exception as e:
        return f"[XLSX Error: {str(e)}]"


# ─── Text Extraction (main dispatcher) ───
def extract_text(filepath, filename):
    """Extract text from any supported format. Returns str."""
    ext = get_ext(filename)

    # Text-based formats
    if ext in ('.txt', '.md', '.csv', '.json', '.py', '.html', '.xml', '.log'):
        for enc in ['utf-8', 'gbk', 'latin-1']:
            try:
                with open(filepath, 'r', encoding=enc) as f:
                    return f.read()
            except (UnicodeDecodeError, UnicodeError):
                continue
        return ""

    # PDF
    if ext == '.pdf':
        text = ""
        try:
            import pdfplumber
            with pdfplumber.open(filepath) as pdf:
                for i, page in enumerate(pdf.pages):
                    t = page.extract_text()
                    if t:
                        text += f"\n--- Page {i+1} ---\n{t}\n"
            if text.strip():
                return text
        except Exception:
            pass
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(filepath)
            for i, page in enumerate(reader.pages):
                t = page.extract_text()
                if t:
                    text += f"\n--- Page {i+1} ---\n{t}\n"
            return text
        except Exception:
            return f"[ERROR: Cannot extract text from {filename}]"

    # DOCX
    if ext == '.docx':
        try:
            from docx import Document
            doc = Document(filepath)
            return "\n".join(p.text for p in doc.paragraphs)
        except Exception:
            return f"[ERROR: Cannot extract text from {filename}]"

    # XLSX / XLS
    if ext in ('.xlsx', '.xls'):
        return extract_xlsx(filepath, filename)

    # Audio / Video — transcribe via ASR
    if ext in MEDIA_EXTS or ext in VIDEO_EXTS:
        return transcribe_media(filepath, filename)

    # Archive — handled at higher level
    if ext in ARCHIVE_EXTS:
        return f"[ARCHIVE: {filename}]"

    return f"[UNSUPPORTED: {ext}]"


# ─── Chunking (unchanged) ───
def chunk_text(text, chunk_size=1000, overlap=200):
    """Semantic chunking: split by headings, paragraphs, then sentences."""
    if not text or not text.strip():
        return []

    heading_pattern = re.compile(r'(^#{1,4}\s+.+$)', re.MULTILINE)
    sections = heading_pattern.split(text)

    raw_chunks = []
    current_section = ""
    current_heading = ""

    for part in sections:
        if heading_pattern.match(part):
            if current_section.strip():
                raw_chunks.append((current_heading, current_section.strip()))
            current_heading = part.strip()
            current_section = ""
        else:
            current_section += part

    if current_section.strip():
        raw_chunks.append((current_heading, current_section.strip()))

    if not raw_chunks:
        raw_chunks = [("", text.strip())]

    result = []
    for heading, section_text in raw_chunks:
        paragraphs = re.split(r'\n\s*\n', section_text)
        current = ""
        for p in paragraphs:
            p = p.strip()
            if not p:
                continue
            if len(current) + len(p) < chunk_size:
                current += p + "\n\n"
            else:
                if current.strip():
                    result.append({"text": current.strip(), "metadata": {"section": heading} if heading else {}})
                if len(p) > chunk_size:
                    sentences = re.split(r'(?<=[。！？.!?])\s*', p)
                    sub = ""
                    for s in sentences:
                        if len(sub) + len(s) < chunk_size:
                            sub += s
                        else:
                            if sub.strip():
                                result.append({"text": sub.strip(), "metadata": {"section": heading} if heading else {}})
                            sub = s
                    if sub.strip():
                        current = sub + "\n\n"
                    else:
                        current = ""
                else:
                    current = p + "\n\n"
        if current.strip():
            result.append({"text": current.strip(), "metadata": {"section": heading} if heading else {}})

    if overlap > 0 and len(result) > 1:
        for i in range(1, len(result)):
            prev_text = result[i-1]["text"]
            if len(prev_text) > overlap:
                result[i]["text"] = prev_text[-overlap:] + "\n\n" + result[i]["text"]

    return [r for r in result if len(r["text"]) > 20]


def process_text(text, title):
    """Process raw text: chunk and return structured data."""
    if not text or not text.strip():
        return {"title": title, "full_text": "", "chunks": []}
    chunks = chunk_text(text)
    return {"title": title, "full_text": text, "chunks": chunks, "chunk_count": len(chunks)}


def process_file(filepath, filename):
    """Process a file: extract text, chunk, return structured data."""
    text = extract_text(filepath, filename)
    title = os.path.splitext(filename)[0]
    return process_text(text, title)


def summarize(text, max_len=200):
    """Extract first meaningful sentences as a summary."""
    if not text:
        return ""
    text = text.strip()
    sentences = re.split(r'(?<=[。！？.!?])\s*', text)
    summary = ""
    for s in sentences:
        s = s.strip()
        if len(s) < 10:
            continue
        if len(summary) + len(s) > max_len:
            break
        summary += s
    return summary if summary else text[:max_len]


if __name__ == "__main__":
    print("✅ document_processor.py supports: PDF, DOCX, XLSX, ZIP, TAR, MP3, MP4 + more")
